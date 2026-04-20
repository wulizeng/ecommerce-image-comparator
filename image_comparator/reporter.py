import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from comparator import CompareResult

_RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_HEADERS = ["url1", "url2", "is_same", "similarity_score", "recommendation", "reason", "method"]


def generate_excel_bytes(results: list[CompareResult]) -> bytes:
    """
    将比对结果列表生成 Excel 文件字节。
    不同图片（is_same=False）的行用红色背景标注。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "比对结果"

    ws.append(_HEADERS)

    for result in results:
        row = [
            result.url1,
            result.url2,
            "是" if result.is_same else "否",
            result.similarity_score,
            result.recommendation,
            result.reason,
            result.method,
        ]
        ws.append(row)
        if not result.is_same:
            for col in range(1, len(_HEADERS) + 1):
                ws.cell(ws.max_row, col).fill = _RED_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
