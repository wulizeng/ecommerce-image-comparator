import io
import pytest
import openpyxl
from comparator import CompareResult
from reporter import generate_excel_bytes


def make_result(is_same: bool) -> CompareResult:
    return CompareResult(
        url1="https://a.com/1.jpg",
        url2="https://a.com/2.jpg",
        is_same=is_same,
        similarity_score=95 if is_same else 10,
        recommendation="图片一致，无需调整" if is_same else "图片不一致，建议调整店铺链接主图",
        reason="测试原因",
        method="phash_same" if is_same else "phash_diff",
    )


def test_generate_excel_bytes_returns_bytes():
    results = [make_result(True), make_result(False)]
    data = generate_excel_bytes(results)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_generate_excel_has_correct_columns():
    results = [make_result(True)]
    data = generate_excel_bytes(results)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, 8)]
    assert headers == ["url1", "url2", "is_same", "similarity_score", "recommendation", "reason", "method"]


def test_generate_excel_row_count():
    results = [make_result(True), make_result(False), make_result(True)]
    data = generate_excel_bytes(results)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    # 1行标题 + 3行数据
    assert ws.max_row == 4


def test_generate_excel_different_rows_have_red_fill():
    """is_same=False 的行应有红色背景"""
    results = [make_result(True), make_result(False)]
    data = generate_excel_bytes(results)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    # 第2行(is_same=True)：无红色填充
    cell_same = ws.cell(2, 1)
    # 第3行(is_same=False)：红色填充
    cell_diff = ws.cell(3, 1)
    assert cell_diff.fill.fgColor.rgb == "FFCCCCCC" or "FFCCCC" in (cell_diff.fill.fgColor.rgb or "")
